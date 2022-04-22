import json, os, argparse
import openpyxl
from openpyxl.styles import  Font, PatternFill, Border, Side, Alignment, Color
import numpy as np


class COMPARE_DSPF():

    def __init__(self, file_path_1, file_path_2, output_path):
        self.filePath1 = file_path_1
        self.filePath2 = file_path_2
        self.outputPath = output_path
        pass
        
        self.currentOperation = None

        self.openFiles()

        self.layerMapDict1 = self.createLayerMapDict(self.fileContentList1)
        jsonOb = json.dumps(self.layerMapDict1, sort_keys=True)
        with open('json1.json', 'w') as wf:
            wf.writelines(jsonOb)

        self.layerMapDict2 = self.createLayerMapDict(self.fileContentList1)
        jsonOb = json.dumps(self.layerMapDict2, sort_keys=True)
        with open('json2.json', 'w') as wf:
            wf.writelines(jsonOb)

        self.createLayerOutput()

    
    def openFiles(self):
        with open(self.filePath1, 'r') as rf:
            self.fileContentList1 = rf.readlines()

        with open(self.filePath2, 'r') as rf:
            self.fileContentList2 = rf.readlines()


    def createLayerMapDict(self, fileContents = []):
        lineNumber = 1
        layerMapDict = {}
        rcContents = []
        for cont in fileContents:
            line = str(cont).strip()
            wordList = line.split() 
            wordLength = len(wordList)
            if wordLength > 0:

                if line.lower() == "* instance section":
                    self.currentOperation = None
                
                if self.currentOperation == 'net':
                    rcContents.append(line)

                if line.upper() == '*LAYER_MAP':
                    self.currentOperation = 'layer'
                    # print(lineNumber, "\t", self.currentOperation)

                elif wordList[0].upper() == '*|NET':
                    self.currentOperation = 'net'


                if self.currentOperation == 'layer' and wordLength > 2 and wordList[0].startswith('*'):
                    tempDict = {}               
                    tempDict['id'] = wordList[0].split('*')[-1]
                    tempDict['name'] =  wordList[1]
                    tempDict['itf'] =  wordList[2].split('=')[-1]
                    tempDict['type'] = {}
    
                    layerMapDict[tempDict['id']] = tempDict

                    

            lineNumber += 1
        
        # print(layerMapDict['632'])
        layerMapDict = self.createRCmapDict(fileContents=rcContents, layerMapDict=layerMapDict)
        
        return layerMapDict

    def createRCmapDict(self,  layerMapDict, fileContents = []):
        tempDict = {}
        # prevNetVal = 1
        # currentNetVal = 1 
        print("createRCmapDict\n\n")
        for cont in fileContents:
            line = str(cont).strip()
            wordList = line.split() 
            wordLength = len(wordList)

            if wordLength > 3:
                if wordList[0].upper().startswith('R'):
                    labelId = ''
                    rvalue = None
                    rName = wordList[0].split("_")[0]
                    for val in wordList:
                        if val.startswith('R=') or val.startswith('r') :
                            rvalue = val.split('=')[-1]
                    if rvalue == None:
                        rvalue = wordList[3]

                    if  wordList[-1].split("=")[0].__contains__('$lvl'):
                        layer = wordList[-1].split("=")[-1] 
                        labelId += layer + "_"             
                    if  wordList[-2].split("=")[0].__contains__('$lvl'):
                        layer = wordList[-2].split("=")[-1] 
                        labelId +=   layer

                    if labelId not in layerMapDict.keys():
                        tempDict = {}               
                        tempDict['id'] = labelId
                        tempDict['name'] =  labelId
                        tempDict['itf'] =  ''
                        tempDict['type'] = {}
                        layerMapDict[tempDict['id']] = tempDict   

                    if rName in layerMapDict[labelId]['type'].keys():
                        layerMapDict[labelId]['type'][rName].append(rvalue)
                    else:
                        layerMapDict[labelId]['type'][rName] = []
                        layerMapDict[labelId]['type'][rName].append(rvalue)


                elif wordList[0].upper().startswith('C'):
                    cValue =  wordList[3]
                    cName = wordList[0].split("_")[0]
                    labelId = ''
                    if  wordList[-2].split("=")[0].__contains__('$lvl'):
                        layer = wordList[-2].split("=")[-1]
                        labelId += layer + "_"
                    if  wordList[-1].split("=")[0].__contains__('$lvl'):
                        layer = wordList[-1].split("=")[-1]
                        labelId +=  layer

                    if labelId not in layerMapDict.keys():
                        tempDict = {}               
                        tempDict['id'] = labelId
                        tempDict['name'] =  labelId
                        tempDict['itf'] =  ''
                        tempDict['type'] = {}
                        layerMapDict[tempDict['id']] = tempDict

                    if cName not in layerMapDict[labelId]['type'].keys():
                        layerMapDict[labelId]['type'][cName] = []
                        layerMapDict[labelId]['type'][cName].append(cValue)
                    else:
                        layerMapDict[labelId]['type'][cName].append(cValue)
                        # print(cName, "\t", cValue)
                        

        # print_dict(layerMapDict)
        return layerMapDict

    def convertToExcelCell(self, ind):
        colInd = ''
        while ind > 0:
            ind, rem =  divmod(ind-1, 26)
            colInd = chr(65 + rem) + colInd
        return colInd

    def getCommonKey(self):

        keyList = sorted(list(self.layerMapDict1.keys()))
        self.layerMapDictMergedKey = []
        tempList = []
        tempDict = {}
        for keyD in keyList:
            if str(keyD).__contains__("_") and keyD in self.layerMapDict2.keys():
                tempKey = int(keyD.split('_')[0])
                if tempKey not in tempList:
                    tempList.append(tempKey)
                    tempDict[tempKey] = []
                
                if keyD not in tempDict[tempKey]:
                    tempDict[tempKey].append(keyD)

        tempList = sorted(tempList)

        print("tempList : \t", tempList)
        # print_dict(tempDict)
        for keyD in tempList:
            for keyL in tempDict[keyD]:
                self.layerMapDictMergedKey.append(keyL)
        
    
    def getUnit(self, val = None):
        unitDict = {
                '-8' : 'y',  # yocto
                '-7' : 'z',  # zepto
                '-6' : 'a',  # atto
                '-5' : 'f',  # femto
                '-4' : 'p',  # pico
                '-3' : 'n',   # nano
                '-2' : 'u',   # micro
                '-1' : 'm',   # mili
                # '-2' : 'c',   # centi
                # '-1' : 'd',   # deci
                '1' : 'k',    # kilo
                '2' : 'M',    # mega
                '3' : 'G',    # giga
                '4' : 'T',   # tera
                '5' : 'P',   # peta
                '6' : 'E',   # exa
                '7' : 'Z',   # zetta
                '8' : 'Y',   # yotta
        }
        unitListKey = list(unitDict.keys())
        if val != None:
            val0 = str(val).strip().lower()
            try:
                valFlt = float(val0)
                valFloat = np.format_float_scientific(valFlt, precision=4, exp_digits=1)
                valFloat = str(valFloat)
                floatVal, floatUnit = valFloat.split('e')
                floatVal = float(floatVal)
                floatUnit = int(floatUnit)
                if floatUnit >= 0 and floatUnit < 3:
                    floatFinalVal = float(floatVal) * (10**int(floatUnit))
                    floatFinalUnit = ''
                    floatFinalVal = round(floatFinalVal, 2)
                    return floatFinalVal, floatFinalUnit
                else:
                    divs = int(floatUnit) // 3
                    rem = int(floatUnit) % 3
                    floatFinalUnit = unitDict[str(divs)]
                    floatFinalVal = float(floatVal) * (10 ** (rem))
                    floatFinalVal = round(floatFinalVal, 2)
                    return floatFinalVal, floatFinalUnit
            except Exception as e:
                print(e)
                valFloat = '0e0'
                return '0', '-'
        return '-'



    def formatColWidth(self, tempWorkSheet, space=2):
        colInd = 1
        for cells in tempWorkSheet.columns:
            maxWidth = int(max(len(str(cell.value).strip()) for cell in cells))
            colName = self.convertToExcelCell(colInd)
            tempWorkSheet.column_dimensions[colName].width = maxWidth + space
            colInd += 1

        return tempWorkSheet  
    

    def addFloatValues(self, valueList):
        tempVal = float(0)
        if type(valueList) == int or type(valueList) == float:
            return valueList

        for val in valueList:
            tempVal += float(val)
        return tempVal

    def insertToCell(self, outputSheet, rowNumber, cellValueStr="", algType='Default'):
        self.valueDelimiter = "<>"
        cellValues = cellValueStr.split(self.valueDelimiter)
        cellStyle = self.cellStyle
        for ind, cellVal in enumerate(cellValues):
            colName = self.convertToExcelCell(ind+1)
            cell = outputSheet[f"{colName}{rowNumber}"] 
            cell.value = cellVal
            cell.alignment = self.cellStyle[algType]['alignment']
            cell.font = cellStyle[algType]['font']


    def createLayerOutput(self):
        self.getCommonKey()
        outputWB = openpyxl.Workbook()
        outputSheet = outputWB.create_sheet('File Comparision')

        self.cellStyle = {
            "Default" : {
            'font': Font(name='Consolas', size=10, bold=False, italic=False, underline='none', strike=False, vertAlign=None),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=False, shrink_to_fit=False, indent=0)
            },

            "Title" : {
            'font': Font(name='Consolas', size=11, bold=True, italic=False, underline='none', strike=False, vertAlign=None),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=False, shrink_to_fit=False, indent=0)
            }
        }
        cellStyle = self.cellStyle

        cell = outputSheet[f"A1"]
        cell.value = "DSPF1"
        cell.alignment = cellStyle['Title']['alignment']
        cell.font = cellStyle['Title']['font']
        cell = outputSheet[f"E1"]
        cell.value = "DSPF2" 
        cell.alignment = cellStyle['Title']['alignment']
        cell.font = cellStyle['Title']['font']
        cell = outputSheet[f"I1"]
        cell.value = "Difference" 
        cell.alignment = cellStyle['Title']['alignment']
        cell.font = cellStyle['Title']['font']
        
        self.insertToCell(outputSheet=outputSheet, rowNumber=2, cellValueStr=f"{'Layer1'}<>{'Layer2'}<>{'Type'}<>{'Value'}<>{'Layer1'}<>{'Layer2'}<>{'Type'}<>{'Value'}<>{'Difference'}", algType='Title')

        outputSheet.merge_cells("A1:D1")
        outputSheet.merge_cells("E1:H1")
        outputSheet.merge_cells("I1:I2")
        

        rowNumber = 3
        print_dict(self.layerMapDictMergedKey)
        for layer in self.layerMapDictMergedKey:
            layer1 = layer.split("_")[0]
            layer2 = layer.split("_")[1]
            if layer2 == '':
                layer2 = 0

            typeDict1 = self.layerMapDict1[layer]['type']
            typeDict2 = self.layerMapDict2[layer]['type']
            # print(typeDict1, "\n", typeDict2, "\n\n\n")
            self.symOhm = '\u03A9'
            cVal1 = float(0)
            cVal2 = float(0)

            rVal1 = float(0)
            rVal2 = float(0)

            print(f"\n\n\n{layer}")
            for keyRC in typeDict1.keys():
                if keyRC in typeDict2.keys():
                    # value1, unit1 = self.getUnit(self.addFloatValues(typeDict1[parasiticType1]))
                    value1 = self.addFloatValues(typeDict1[keyRC])
                    value2 = self.addFloatValues(typeDict2[keyRC])
                    
                    # value2, unit2 = self.getUnit(self.addFloatValues(typeDict2[parasiticType2]))
                    # difference = "N/A"
                    # difference = ((value2-value1)/value2)*100
                    
                    if keyRC.startswith("C"):
                        # value1 = f"{value1} {unit1}F"
                        # value2 = f"{value2} {unit2}F"
                        cVal1 += value1
                        cVal2 += value2
                        currentType = 'C'
                    elif keyRC.startswith("R"):
                        # value1 = f"{value1} {unit1}{self.symOhm}"
                        # value2 = f"{value2} {unit2}{self.symOhm}"
                        rVal1 += value1
                        rVal2 += value2
                        currentType = 'R'
                    
                    else:
                        print("No RC")
                    


                    print(f"\n{layer} - {currentType} : ", value1, value2)
                    print("C : ", cVal1, cVal1)
                    print("R : ", rVal1, rVal2)
                # else:
                #     parasiticType1 = keyRC
                #     value1 = self.addFloatValues(typeDict1[parasiticType1])
                #     if keyRC.startswith("C"):
                #         # value1 = f"{value1} {unit1}F"
                #         # value2 = f"{value2} {unit2}F"
                #         cVal1 += value1
                #         # cVal2 += value2
                #     if keyRC.startswith("R"):
                #         # value1 = f"{value1} {unit1}{self.symOhm}"
                #         # value2 = f"{value2} {unit2}{self.symOhm}"
                #         rVal1 += value1
                #         # rVal2 += value2

                #     # parasiticType2 = "N/A"
                #     # value2 = "N/A"

                #     # difference = "N/A"

            value1, unit1 = self.getUnit(cVal1)
            value2, unit2 = self.getUnit(cVal2)
            
            try:
                cDifference = ((value2-value1)/value2)*100
            except:
                cDifference = "N/A"
            cVal1 = f"{value1} {unit1}{self.symOhm}"
            cVal2 = f"{value2} {unit2}{self.symOhm}"   
            
            if int(value2) != 0:
                valueString = f"{layer1}<>{layer2}<>{'C'}<>{cVal1}<>{layer1}<>{layer2}<>{'C'}<>{cVal2}<>{cDifference}"
                self.insertToCell(outputSheet=outputSheet, rowNumber=rowNumber, cellValueStr=valueString)
                rowNumber += 1

            value1, unit1 = self.getUnit(rVal1)
            value2, unit2 = self.getUnit(rVal2)
            try:
                rDifference = ((value2-value1)/value2)*100
            except:
                rDifference = "N/A"
            rVal1 = f"{value1} {unit1}{self.symOhm}"
            rVal2 = f"{value2} {unit2}{self.symOhm}"
            
            if int(value2) != 0:
                valueString = f"{layer1}<>{layer2}<>{'R'}<>{rVal1}<>{layer1}<>{layer2}<>{'R'}<>{rVal2}<>{rDifference}"
                self.insertToCell(outputSheet=outputSheet, rowNumber=rowNumber, cellValueStr=valueString)
                rowNumber += 1
            

        outputSheet = self.formatColWidth(tempWorkSheet=outputSheet, space=5)
        
        sheetList = outputWB.sheetnames
        del outputWB[sheetList[0]]
        outputXlsxDirectory = self.outputPath

        if os.path.isfile(outputXlsxDirectory):
            os.remove(outputXlsxDirectory)

        outputWB.save(outputXlsxDirectory)
        outputWB.close()





def print_dict(tempDict):
    try:
        jsonObj = json.dumps(tempDict, indent=4)
        print(jsonObj)
    except:
        print(tempDict)


def compareDspf(file_path_1, file_path_2, output_path='.'):
    print("\n\n", file_path_1, file_path_2, output_path, "\n\n")

    compare_dspf_class = COMPARE_DSPF(file_path_1, file_path_2, output_path)


if __name__ == "__main__":
    
    curWorkingDir = str(os.getcwd()).replace("\\", "/")
    command_format = "python {script_Path} {path_of_dspf_file_1} {path_of_dspf_file_2} {output_file_directory}"
    parser = argparse.ArgumentParser(description=command_format)
    parser.add_argument('dspf_1', type=str)
    parser.add_argument('dspf_2', type=str)
    # parser.add_argument('output', type=str, default=".")
    args = parser.parse_args()

    try:
        file_path_1 = args.dspf_1
        file_path_2 = args.dspf_2
        # output_path = args.output
    except:
        print(f"\n\nGive Command As Follows : {command_format}")
    
    output_path = f"{curWorkingDir}/ouput_dspf_compare.xlsx"

    compareDspf(file_path_1, file_path_2, output_path)



# python -u "c:\Users\dtco-gf\Desktop\AUTOMATION_TEAM\DSPF\cboa_experiment\DSPF\22FDX\compare_dspf_files.py" "C:\Users\dtco-gf\Desktop\AUTOMATION_TEAM\DSPF\cboa_experiment\DSPF\22FDX\SC12T_INVX8_CSC20R_CBoA.nominal.dspf" "C:\Users\dtco-gf\Desktop\AUTOMATION_TEAM\DSPF\cboa_experiment\DSPF\22FDX\SC12T_INVX8_CSC20R.nominal.dspf" "."



